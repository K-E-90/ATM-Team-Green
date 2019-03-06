from openpyxl import load_workbook
import numpy as np
from math import *
import matplotlib.pyplot as plt

wb_sw = load_workbook("trafficSW-stack-departures.xlsx", read_only=True)
wb_ne = load_workbook("trafficNE-stack-departures.xlsx", read_only=True)
traffic_SW_dep = wb_sw["trafficSW-stack-departures"]
traffic_NE_dep = wb_ne["trafficNE-stack-departures"]

traffic_sw_dep = np.array([[i.value for i in j] for j in traffic_SW_dep['A2':'H132']])
traffic_ne_dep = np.array([[i.value for i in j] for j in traffic_NE_dep['A2':'H132']])

airport = 'EHAM'

def graph_dep(traffic, airport):
    direction = np.array([])
    coordinates_lat = np.array([])
    coordinates_lon = np.array([])
    for i in range(len(traffic)):
        if traffic[i,0]==airport:
            direction= np.append(direction,traffic[i,1])
            coordinates_lat = np.append(coordinates_lat,cos(degrees(traffic[i,1]))*230)
            coordinates_lon = np.append(coordinates_lon,sin(degrees(traffic[i,1]))*230)
    return coordinates_lat, coordinates_lon

lon_sw = graph_dep(traffic_sw_dep,airport)[0]
lat_sw = graph_dep(traffic_sw_dep,airport)[1]
lon_ne = graph_dep(traffic_ne_dep,airport)[0]
lat_ne = graph_dep(traffic_ne_dep,airport)[1]
plt.figure(1)
plt.plot(lon_ne, lat_ne, 'ro')
plt.grid(True)
plt.show()
        
        